# -*- coding: utf-8 -*-

import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from glob import glob
from multiprocessing import cpu_count

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils import DATA_DIR as BASE_DATA_DIR
from utils import OUT_DIR as BASE_OUT_DIR
from utils import create_logger, log_filename, normalize_value, touch

DATA_DIR = os.path.join(BASE_DATA_DIR, 'desa')
OUT_DIR = os.path.join(BASE_OUT_DIR, 'desa')

os.makedirs(OUT_DIR, exist_ok=True)

logger = create_logger(log_filename(__file__))
patches = {
    '12.03.04.2064': 'Pangurabaan',
    '12.14.06.2011': 'Hiliganowo',
    '12.20.07.2004': 'Pintu Padang',
    '14.01.02.2030': 'Naumbai',
    '14.01.07.2030': 'Kuntu Darussalam',
    '14.02.03.2022': 'Dusun Tua',
    '14.07.05.2025': 'Bakti Makmur',
    '14.09.01.2015': 'Pulaubinjai',
    '14.10.05.2004': 'Bagan Melibur',
    '72.01.02.2017': 'Laonggo',
    '72.01.02.2028': 'Nanga-Nangaon',
    '72.07.06.2021': 'Meselesek',
}
excludes = {
    '13.02.02.1036'
}


def lookup_name(ws: Worksheet, min_row: int, name_ncol: int) -> str:
    nrow = min_row + 2
    name = ws.cell(nrow, name_ncol).value
    if name is None:
        name = ws.cell(nrow, name_ncol+1).value
    return name


def extract_data(src_path: str):
    logger.info('Processing %s', src_path)
    basename, ext = os.path.splitext(os.path.basename(src_path))
    dest_path = os.path.join(OUT_DIR, '{}-out{}'.format(basename, ext))

    try:
        touch(dest_path)
    except Exception as e:
        logger.error('Could not write output file: %s', dest_path)
        logger.error('%s', e)
        return

    wb = load_workbook(src_path, read_only=True)
    ws = wb.active

    list_kode = []
    list_nama = []

    for row in ws.iter_rows(ws.min_row, ws.max_row, 1, 7, True):
        kode, v2, v3, v4, v5, v6, v7 = row
        if kode is not None:
            kode = str(kode).strip()
            for k in kode.split('\n'):
                k = k.strip()
                if re.match(r'^(\d{2}\.){3}\d{4}$', k):
                    list_kode.append(k)
                    if k in patches:
                        list_nama.append(patches[k])

        cols_nama = [v6, v7]
        if v6 is None and v7 is None:
            cols_nama.append(v5)
        for nama in cols_nama:
            if nama is None:
                continue
            nama = str(nama).strip()
            if nama.startswith('6 Bambalemo'):
                list_nama.append('Bambalemo Ranomaisi')
                continue
            for n in nama.split('\n'):
                if re.match(r'^\d+ .+$', n.strip()):
                    val = normalize_value(n)
                    if os.path.basename(src_path) == '6.xlsx' and val in ['Ulu', 'Ilir']:
                        continue
                    list_nama.append(val)

    wb.close()

    count_kode = len(list_kode)
    count_nama = len(list_nama)
    if count_kode == count_nama:
        logger.info('Kode: %s. Nama: %s. %s', count_kode, count_nama, dest_path)
    else:
        logger.warn('Kode: %s. Nama: %s. %s', count_kode, count_nama, dest_path)

    if count_nama > count_kode:
        list_kode.extend([None for i in range(count_nama-count_kode)])
    elif count_kode > count_nama:
        list_nama.extend([None for i in range(count_kode-count_nama)])

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title='DESA-KELURAHAN')
    ws_out.column_dimensions['A'].width = 14
    ws_out.column_dimensions['B'].width = 50

    for row in zip(list_kode, list_nama):
        if row[0] in excludes:
            continue
        ws_out.append(row)

    try:
        wb_out.save(dest_path)
        logger.info('Data succesfully saved to: %s', dest_path)
    except Exception as e:
        logger.error('Could not save to: %s', dest_path)
        logger.error('%s', e)


def join_files():
    dest_path = os.path.join(BASE_OUT_DIR, 'desa-out.xlsx')
    try:
        touch(dest_path)
    except Exception as e:
        logger.error('Could not write output file: %s', dest_path)
        logger.error('%s', e)
        return

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet('DESA-KELURAHAN')
    ws_out.column_dimensions['A'].width = 14
    ws_out.column_dimensions['B'].width = 50

    fpaths = glob(os.path.join(OUT_DIR, '*.xlsx'))
    for fpath in fpaths:
        wb = load_workbook(fpath, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(ws.min_row, ws.max_row, 1, 2, True):
            v1, v2 = row
            if v1 is None or v2 is None:
                logger.warn('Empty row in %s: %s, %s', fpath, v1, v2)
            ws_out.append(row)
        wb.close()

    try:
        wb_out.save(dest_path)
        logger.info('Data succesfully saved to: %s', dest_path)
    except Exception as e:
        logger.error('Could not save to: %s', dest_path)
        logger.error('%s', e)


if __name__ == '__main__':
    t0 = time.process_time()
    fnames = glob(os.path.join(DATA_DIR, '*.xlsx'))
    with ThreadPoolExecutor(max_workers=cpu_count()) as executor:
        executor.map(extract_data, fnames)
        try:
            executor.shutdown(wait=True)
        except Exception as e:
            logger.error('%s', e)

    join_files()
    t1 = time.process_time()
    td = round(t1-t0, 4)
    logger.info('Elapsed time: %s seconds', td)
