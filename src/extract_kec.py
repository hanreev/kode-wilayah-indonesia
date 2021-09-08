# -*- coding: utf-8 -*-

import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from multiprocessing import cpu_count
from glob import glob
from openpyxl import Workbook, load_workbook

from utils import DATA_DIR, OUT_DIR, create_logger, log_filename, normalize_value, touch

logger = create_logger(log_filename(__file__))


def get_output_path(src_path: str) -> str:
    basename, ext = os.path.splitext(os.path.basename(src_path))
    return os.path.join(OUT_DIR, '{}-out{}'.format(basename, ext))


def extract_data(src_path):
    logger.info('Processing %s', src_path)

    dest_path = get_output_path(src_path)

    try:
        touch(dest_path)
    except Exception as e:
        logger.error('Could not write output file: %s', dest_path)
        logger.error('%s', e)
        return

    wb = load_workbook(src_path, read_only=True)
    ws = wb.active

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title='KECAMATAN')
    ws_out.column_dimensions['A'].width = 9
    ws_out.column_dimensions['B'].width = 50

    for row in ws.iter_rows(ws.min_row, ws.max_row, 1, 4, True):
        v1, v2, v3, v4 = row
        if v1 is None and v2 is None:
            continue
        if v2 is not None:
            kode = str(v2).strip()
            nama = str(v3 or v4).strip()
        else:
            kode = str(v1).strip()
            nama = str(v2 or v3).strip()
        if not re.match(r'^(\d{2}\.){2}\d{2}$', kode):
            continue
        ws_out.append([kode, normalize_value(nama)])

    wb.close()

    try:
        wb_out.save(dest_path)
        logger.info('Data succesfully saved to: %s', dest_path)
    except Exception as e:
        logger.error('Could not save to: %s', dest_path)
        logger.error('%s', e)


def join_files(fpaths):
    dest_path = os.path.join(OUT_DIR, 'kecamatan-out.xlsx')

    try:
        touch(dest_path)
    except Exception as e:
        logger.error('Could not write output file: %s', dest_path)
        logger.error('%s', e)
        return

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title='KECAMATAN')
    ws_out.column_dimensions['A'].width = 9
    ws_out.column_dimensions['B'].width = 50

    for fpath in fpaths:
        opath = get_output_path(fpath)
        wb = load_workbook(opath, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            ws_out.append(row)
        wb.close()

    try:
        wb_out.save(dest_path)
        logger.info('Data succesfully saved to: %s', dest_path)
    except Exception as e:
        logger.error('Could not save to: %s', dest_path)
        logger.error('%s', e)


if __name__ == '__main__':
    t0 = time.process_time()
    fnames = glob(os.path.join(DATA_DIR, 'kecamatan*.xlsx'))
    with ThreadPoolExecutor(max_workers=cpu_count()) as executor:
        executor.map(extract_data, fnames)
        try:
            executor.shutdown(wait=True)
        except Exception as e:
            logger.error('%s', e)

    join_files(fnames)
    t1 = time.process_time()
    td = round(t1-t0, 4)
    logger.info('Elapsed time: %s seconds', td)
