# -*- coding: utf-8 -*-

import os
import re
import sqlite3
import time
from subprocess import PIPE, Popen

import mysql.connector
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from utils import BASE_DIR, DATA_DIR, OUT_DIR, create_logger, log_filename, touch

OUTPUT_BASENAME = 'kode_wilayah_indonesia'
DB_CONFIG = {
    'postgres': {
        'host': 'localhost',
        'port': '5432',
        'user': 'postgres',
        'password': '1'
    },
    'mysql': {
        'host': 'localhost',
        'port': '3306',
        'user': 'root',
        'password': '1'
    }
}

logger = create_logger(log_filename(__file__))


def main():
    dest_path = os.path.join(BASE_DIR, '{}.xlsx'.format(OUTPUT_BASENAME))
    try:
        touch(dest_path)
    except Exception as e:
        logger.error('Could not write output file: %s', dest_path)
        logger.error('%s', e)
        return

    src_dict = {
        'PROVINSI': (os.path.join(DATA_DIR, 'provinsi.xlsx'), ('KODE', 'PROVINSI')),
        'KABUPATEN/KOTA': (os.path.join(OUT_DIR, 'kabupaten-out.xlsx'), ('KODE PROVINSI', 'KODE', 'KABUPATEN/KOTA')),
        'KECAMATAN': (os.path.join(OUT_DIR, 'kecamatan-out.xlsx'), ('KODE KABUPATEN/KOTA', 'KODE', 'KECAMATAN')),
        'DESA/KELURAHAN': (os.path.join(OUT_DIR, 'desa-out.xlsx'), ('KODE KECAMATAN', 'KODE', 'DESA/KELURAHAN')),
    }

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for ws_name, (fpath, headers) in src_dict.items():
        ws_out = wb_out.create_sheet(title=ws_name.replace('/', '-'))
        ws_out.append(headers)
        ws_out.column_dimensions['A'].auto_size = True
        ws_out.column_dimensions['B'].auto_size = True
        ws_out.column_dimensions['C'].auto_size = True
        wb = load_workbook(fpath, read_only=True)
        ws = wb.active
        nrow = 1
        for row in ws.iter_rows(ws.min_row, ws.max_row, 1, 2, True):
            v1, v2 = row
            if v1 is None or v2 is None:
                logger.warn('Empty row in %s: %s, %s', fpath, v1, v2)
            if ws_name == 'PROVINSI':
                ws_out.append(row)
            else:
                fk = re.sub(r'\.\d+$', '', str(v1))
                ws_out.append([fk, v1, v2])
            nrow += 1
        wb.close()
        ncol = 1
        for col in ws_out.iter_cols(1, 2, 1, 2, True):
            width = max(len(str(val)) for val in col) + 4
            ws_out.column_dimensions[get_column_letter(ncol)].width = width
            ncol += 1
        ws_out.column_dimensions['B' if ws_name == 'PROVINSI' else 'C'].width = 50
        table_ref = 'A1:B{}'.format(nrow) if ws_name == 'PROVINSI' else 'A1:C{}'.format(nrow)
        table = Table(displayName=ws_name.replace('/', '_'), ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name='TableStyleMedium9', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws_out.add_table(table)

    try:
        wb_out.save(dest_path)
        logger.info('Data succesfully saved to: %s', dest_path)
    except Exception as e:
        logger.error('Could not save to: %s', dest_path)
        logger.error('%s', e)


def insert_into_db(dbms='postgres', dbname=OUTPUT_BASENAME):
    kwargs = DB_CONFIG.get(dbms)
    if dbms == 'postgres':
        kwargs.update({'dbname': dbname})
        conn = psycopg2.connect(**kwargs)
    elif dbms == 'mysql':
        kwargs.update({'database': dbname})
        conn = mysql.connector.connect(**kwargs)
    elif dbms == 'sqlite':
        conn = sqlite3.connect(os.path.join(BASE_DIR, '{}.db'.format(dbname)))
    else:
        raise ValueError('Invalid DBMS')

    cur = conn.cursor()

    table_dict = {
        'provinsi': [
            'kode VARCHAR(20) NOT NULL PRIMARY KEY',
            'provinsi VARCHAR(255) NOT NULL',
        ],
        'kabupaten_kota': [
            'kode VARCHAR(20) NOT NULL PRIMARY KEY',
            'kode_provinsi VARCHAR(20) NOT NULL',
            'kabupaten_kota VARCHAR(255) NOT NULL',
            'CONSTRAINT kabupaten_kota_kode_provinsi_fk FOREIGN KEY(kode_provinsi) REFERENCES provinsi(kode)',
        ],
        'kecamatan': [
            'kode VARCHAR(20) NOT NULL PRIMARY KEY',
            'kode_kabupaten_kota VARCHAR(20) NOT NULL',
            'kecamatan VARCHAR(255) NOT NULL',
            'CONSTRAINT kecamatan_kode_kabupaten_kota_fk FOREIGN KEY(kode_kabupaten_kota) REFERENCES kabupaten_kota(kode)',
        ],
        'desa_kelurahan': [
            'kode VARCHAR(20) NOT NULL PRIMARY KEY',
            'kode_kecamatan VARCHAR(20) NOT NULL',
            'desa_kelurahan VARCHAR(255) NOT NULL',
            'CONSTRAINT desa_kelurahan_kode_kecamatan_fk FOREIGN KEY(kode_kecamatan) REFERENCES kecamatan(kode)',
        ]
    }

    src_dict = {
        'provinsi': os.path.join(DATA_DIR, 'provinsi.xlsx'),
        'kabupaten_kota': os.path.join(OUT_DIR, 'kabupaten-out.xlsx'),
        'kecamatan': os.path.join(OUT_DIR, 'kecamatan-out.xlsx'),
        'desa_kelurahan': os.path.join(OUT_DIR, 'desa-out.xlsx'),
    }

    # Disable Foreign Key checks for MySQL
    if dbms == 'mysql':
        cur.execute('SET FOREIGN_KEY_CHECKS = 0')

    for table_name, fpath in src_dict.items():
        drop_sql = 'DROP TABLE IF EXISTS {}'.format(table_name)
        if dbms != 'sqlite':
            drop_sql += ' CASCADE'
        cur.execute(drop_sql)
        cur.execute('CREATE TABLE {} ({})'.format(table_name, ','.join(table_dict[table_name])))
        wb = load_workbook(fpath, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(ws.min_row, ws.max_row, 1, 2, True):
            v1, v2 = row
            if v1 is None or v2 is None:
                logger.warn('Empty row in %s: %s, %s', fpath, v1, v2)
            if table_name == 'provinsi':
                sql = 'INSERT INTO {} VALUES  (%s, %s)'.format(table_name)
                args = row
            else:
                fk = re.sub(r'\.\d+$', '', str(v1))
                sql = 'INSERT INTO {} VALUES  (%s, %s, %s)'.format(table_name)
                args = (v1, fk, v2)
            if dbms == 'sqlite':
                sql = sql.replace('%s', '?')
            cur.execute(sql, args)

        wb.close()

    # Re-enable Foreign Key checks for MySQL
    if dbms == 'mysql':
        cur.execute('SET FOREIGN_KEY_CHECKS = 1')

    conn.commit()
    cur.close()
    conn.close()


def dump_db(dbms='postgres', dbname=OUTPUT_BASENAME):
    env = None
    config = DB_CONFIG.get(dbms)
    if dbms == 'postgres':
        cmds = ['pg_dump', '-h', config['host'], '-p', config['port'], '-U',
                config['user'], '-CcOx', '-Fc', '-f{}.backup'.format(dbname), dbname]
        env = {'PGPASSWORD': config['password']}
    elif dbms == 'mysql':
        cmds = ['mysqldump', '-h', config['host'], '-P', config['port'], '-u',
                config['user'], '-p'+config['password'], '-B', dbname, '-r{}.sql'.format(dbname)]
    else:
        raise ValueError('Invalid DBMS')

    logger.info('Executing: %s', ' '.join(cmds))

    with Popen(cmds, env=env, stdout=PIPE, stderr=PIPE) as proc:
        stdout, stderr = proc.communicate()
        if stdout:
            logger.info('%s', stdout.decode('utf-8'))
        if stderr:
            logger.error('%s', stderr.decode('utf-8'))
        retval = proc.wait()
    logger.info('Process exited with return code %s', retval)


if __name__ == '__main__':
    t0 = time.process_time()
    main()
    # insert_into_db(dbms='postgres')
    # insert_into_db(dbms='mysql')
    # insert_into_db(dbms='sqlite')
    # dump_db(dbms='postgres')
    # dump_db(dbms='mysql')
    t1 = time.process_time()
    td = round(t1-t0, 4)
    logger.info('Elapsed time: %s seconds', td)
