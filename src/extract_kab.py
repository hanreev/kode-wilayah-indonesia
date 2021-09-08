# -*- coding: utf-8 -*-

import os
import re
import time

from openpyxl import Workbook, load_workbook

from utils import DATA_DIR, OUT_DIR, create_logger, log_filename, touch

logger = create_logger(log_filename(__file__))


def main():
    src_path = os.path.join(DATA_DIR, 'kabupaten.xlsx')
    dest_path = os.path.join(OUT_DIR, 'kabupaten-out.xlsx')

    logger.info('Processing %s', src_path)

    try:
        touch(dest_path)
    except Exception as e:
        logger.error('Could not write output file: %s', dest_path)
        logger.error('%s', e)
        return

    wb = load_workbook(src_path, read_only=True)
    ws = wb.active

    list_kode = []
    list_nama = []
    dict_kode = {}
    dict_nama = {}

    nrow = 0

    for row in ws.iter_rows(1, ws.max_row, 2, 3, True):
        nrow += 1
        kode, nama = row
        if kode is None:
            continue
        kode = str(kode).strip()
        if len(kode) < 5:
            kode += '0'
        if not re.match(r'^\d{2}\.\d{2}$', kode):
            continue
        list_kode.append(kode)
        dict_kode[nrow] = kode
        if nama is None:
            continue
        nama = re.sub(r'(\w) (KAB|KAB\.|KOTA) (?!KAB |KAB\. |KOTA )', r'\1\n\2 ', nama.strip())
        for i, n in enumerate(nama.split('\n')):
            dict_nama[nrow+i] = re.sub(r'^(KAB) ', r'\1. ', n)

    wb.close()

    for nrow, nama in dict_nama.items():
        if nrow in dict_kode:
            list_nama.append(nama)

    logger.info('Kode: %s. Nama: %s', len(list_kode), len(list_nama))

    wb_out = Workbook(write_only=True)
    ws_out = wb_out.create_sheet(title='KABUPATEN-KOTA')
    ws_out.column_dimensions['A'].width = 6
    ws_out.column_dimensions['B'].width = 50

    for row in zip(list_kode, list_nama):
        ws_out.append(row)

    try:
        wb_out.save(dest_path)
        logger.info('Data succesfully saved to: %s', dest_path)
    except Exception as e:
        logger.error('Could not save to: %s', dest_path)
        logger.error('%s', e)


if __name__ == '__main__':
    t0 = time.process_time()
    main()
    t1 = time.process_time()
    td = round(t1-t0, 4)
    logger.info('Elapsed time: %s seconds', td)
